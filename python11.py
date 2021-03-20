import pandas as pd
from openpyxl import load_workbook
h = int(input('Enter your ps no:-'))
name = str(input('Enter the Name:-'))
email = str(input('Enter the email:-'))
z = pd.read_excel('pythonexcel2.xlsx', sheet_name=['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5'])
y = z['Sheet1']
y = y[(y['PS number'] == h) & (y['Display Name'] == name) & (y['Official Email Address'] == email)]
df = pd.DataFrame(y, columns=['SL#', 'PS number', 'Display Name', 'Official Email Address'])
for i in z.keys():
    x = z[i]
    t = x[(x['PS number'] == h) & (x['Display Name'] == name) & (x['Official Email Address'] == email)]
    col = x.columns
    for j in col:
        df[j] = t[j]




b = load_workbook('pythonexcel2.xlsx')
w = pd.ExcelWriter('pythonexcel2.xlsx', engine='openpyxl')
w.b = b
if 'MasterSheet' in b.sheetnames:
    ref = b['Master Sheet']
    # removing the previous data in the sheet
    b.remove(ref)
df.to_excel(w, sheet_name='Master Sheet')
w.save()
w.close()
"""
Enter your ps no:-99003784
Enter the Name:-Souvik Jana
Enter the email:-souvik.jana@ltts.com
Enter your ps no:-99003786
Enter the Name:-Sangam Patel
Enter the email:-sangam.patel@ltts.com
"""